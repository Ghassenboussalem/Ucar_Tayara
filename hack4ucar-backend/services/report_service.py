from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor, white, black
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
from datetime import datetime

# ─── Arabic reshaping setup ────────────────────────────────────────────────────
_ARABIC_FONT = "Helvetica"  # fallback

try:
    import arabic_reshaper
    from bidi.algorithm import get_display as bidi_display

    _TRADO = r"C:\Windows\Fonts\trado.ttf"
    if os.path.exists(_TRADO):
        pdfmetrics.registerFont(TTFont("TraditionalArabic", _TRADO))
        _ARABIC_FONT = "TraditionalArabic"

    def _ar(text: str) -> str:
        reshaped = arabic_reshaper.reshape(text)
        return bidi_display(reshaped)

except Exception:
    def _ar(text: str) -> str:
        return text


# ─── Colours ───────────────────────────────────────────────────────────────────
UCAR_BLUE      = HexColor('#1a3c6e')
UCAR_LIGHT     = HexColor('#e8f0fa')
ALERT_RED      = HexColor('#c0392b')
ALERT_ORANGE   = HexColor('#e67e22')
SUCCESS_GREEN  = HexColor('#27ae60')
ESG_GREEN      = HexColor('#1e8449')
ESG_LIGHT      = HexColor('#d5f5e3')
RESEARCH_PURPLE = HexColor('#6c3483')
RESEARCH_LIGHT  = HexColor('#f0e6f6')


# ─── Label dictionaries ────────────────────────────────────────────────────────
_L = {
    "fr": {
        # header
        "university":     "UNIVERSITÉ DE CARTHAGE",
        "report_title":   "Rapport Institutionnel",
        "period_label":   "Période",
        "generated":      "Généré le",
        "code":           "Code",
        "governorate":    "Gouvernorat",
        "director":       "Directeur",
        "period_covered": "Période couverte",
        # sections
        "exec_summary":   "Résumé Exécutif (IA)",
        "academic":       "Indicateurs Académiques",
        "finance":        "Indicateurs Financiers",
        "hr":             "Indicateurs Ressources Humaines",
        "infrastructure": "Indicateurs Infrastructure",
        "partnership":    "Indicateurs Partenariats & Mobilité",
        "employment":     "Indicateurs Employabilité des Diplômés",
        "esg":            "Indicateurs ESG / Développement Durable",
        "research":       "Indicateurs Recherche Scientifique",
        "alerts":         "Alertes Actives",
        # table headers
        "indicator":      "Indicateur",
        "value":          "Valeur",
        "status":         "Statut",
        "severity":       "Sévérité",
        "domain":         "Domaine",
        "title":          "Titre",
        # academic
        "total_enrolled":    "Étudiants inscrits",
        "total_passed":      "Étudiants reçus",
        "success_rate":      "Taux de réussite",
        "dropout_rate":      "Taux d'abandon",
        "attendance_rate":   "Taux de présence",
        "repetition_rate":   "Taux de redoublement",
        "avg_grade":         "Note moyenne",
        # finance
        "allocated_budget":         "Budget alloué",
        "consumed_budget":          "Budget consommé",
        "budget_execution_rate":    "Taux d'exécution",
        "cost_per_student":         "Coût par étudiant",
        "staff_budget_pct":         "% Budget RH",
        "infrastructure_budget_pct": "% Budget Infrastructure",
        "research_budget_pct":      "% Budget Recherche",
        # hr
        "total_teaching_staff":    "Personnel enseignant",
        "total_admin_staff":       "Personnel administratif",
        "absenteeism_rate":        "Taux d'absentéisme",
        "avg_teaching_load_hours": "Charge horaire moy.",
        "staff_turnover_rate":     "Taux de rotation",
        "training_completion_rate": "Taux de formation",
        "permanent_staff_pct":     "% Permanent",
        "contract_staff_pct":      "% Contractuel",
        # infrastructure
        "classroom_occupancy_rate":    "Occupation des salles",
        "it_equipment_status_pct":     "Équipements IT (%)",
        "equipment_availability_rate": "Disponibilité équipements",
        "lab_availability_rate":       "Disponibilité labo",
        "library_capacity_used_pct":   "Occupation bibliothèque",
        "ongoing_works":               "Travaux en cours",
        "maintenance_requests":        "Demandes maintenance",
        "resolved_requests":           "Demandes résolues",
        # partnership
        "active_national_agreements":      "Accords nationaux actifs",
        "active_international_agreements": "Accords internationaux actifs",
        "incoming_students":   "Étudiants entrants",
        "outgoing_students":   "Étudiants sortants",
        "erasmus_partnerships": "Partenariats Erasmus+",
        "joint_programs":       "Programmes conjoints",
        "industry_partnerships": "Partenariats industrie",
        "international_projects": "Projets internationaux",
        # employment
        "graduates_total":              "Total diplômés",
        "employed_within_6months":      "Employés à 6 mois",
        "employed_within_12months":     "Employés à 12 mois",
        "employability_rate_6m":        "Taux employabilité 6m",
        "employability_rate_12m":       "Taux employabilité 12m",
        "avg_months_to_employment":     "Délai moyen emploi",
        "national_employment_pct":      "% Emploi national",
        "international_employment_pct": "% Emploi international",
        "self_employed_pct":            "% Auto-entrepreneur",
        # esg
        "energy_consumption_kwh":   "Consommation énergie",
        "carbon_footprint_tons":    "Empreinte carbone",
        "recycling_rate":           "Taux de recyclage",
        "green_spaces_sqm":         "Espaces verts",
        "sustainable_mobility_pct": "Mobilité durable",
        "accessibility_score":      "Score accessibilité",
        "waste_produced_tons":      "Déchets produits",
        "water_consumption_m3":     "Consommation eau",
        # research
        "publications_count":            "Publications",
        "active_projects":               "Projets actifs",
        "funding_secured_tnd":           "Financements obtenus",
        "phd_students":                  "Doctorants inscrits",
        "patents_filed":                 "Brevets déposés",
        "international_collaborations":  "Collaborations internationales",
        "national_collaborations":       "Collaborations nationales",
        "conferences_attended":          "Conférences",
        # footer
        "footer": "Document généré automatiquement par UCAR Intelligence Platform · {n} domaines couverts · Confidentiel · Usage interne uniquement",
        # status
        "normal":    "✅ Normal",
        "attention": "⚠️ Attention",
        "critical":  "🔴 Critique",
        # units
        "unit_pct":    "%",
        "unit_tnd":    " TND",
        "unit_kwh":    " kWh",
        "unit_tons":   " tonnes",
        "unit_sqm":    " m²",
        "unit_m3":     " m³",
        "unit_grade":  "/20",
        "unit_months": " mois",
        "unit_hrs":    "h/sem",
    },
    "ar": {
        "university":     "جامعة قرطاج",
        "report_title":   "التقرير المؤسسي",
        "period_label":   "الفترة",
        "generated":      "أُنشئ في",
        "code":           "الرمز",
        "governorate":    "الولاية",
        "director":       "المدير",
        "period_covered": "الفترة المغطاة",
        "exec_summary":   "الملخص التنفيذي (الذكاء الاصطناعي)",
        "academic":       "المؤشرات الأكاديمية",
        "finance":        "المؤشرات المالية",
        "hr":             "مؤشرات الموارد البشرية",
        "infrastructure": "مؤشرات البنية التحتية",
        "partnership":    "مؤشرات الشراكات والتنقل",
        "employment":     "مؤشرات توظيف الخريجين",
        "esg":            "مؤشرات التنمية المستدامة",
        "research":       "مؤشرات البحث العلمي",
        "alerts":         "التنبيهات النشطة",
        "indicator":      "المؤشر",
        "value":          "القيمة",
        "status":         "الحالة",
        "severity":       "الخطورة",
        "domain":         "المجال",
        "title":          "العنوان",
        "total_enrolled":    "الطلاب المسجلون",
        "total_passed":      "الطلاب الناجحون",
        "success_rate":      "نسبة النجاح",
        "dropout_rate":      "نسبة الانقطاع",
        "attendance_rate":   "نسبة الحضور",
        "repetition_rate":   "نسبة الإعادة",
        "avg_grade":         "المعدل العام",
        "allocated_budget":         "الميزانية المخصصة",
        "consumed_budget":          "الميزانية المستهلكة",
        "budget_execution_rate":    "نسبة تنفيذ الميزانية",
        "cost_per_student":         "تكلفة الطالب",
        "staff_budget_pct":         "% ميزانية الموارد البشرية",
        "infrastructure_budget_pct": "% ميزانية البنية التحتية",
        "research_budget_pct":      "% ميزانية البحث",
        "total_teaching_staff":    "الكادر التدريسي",
        "total_admin_staff":       "الكادر الإداري",
        "absenteeism_rate":        "نسبة الغياب",
        "avg_teaching_load_hours": "متوسط العبء التدريسي",
        "staff_turnover_rate":     "نسبة دوران الموظفين",
        "training_completion_rate": "نسبة إتمام التدريب",
        "permanent_staff_pct":     "% الموظفون الدائمون",
        "contract_staff_pct":      "% الموظفون المتعاقدون",
        "classroom_occupancy_rate":    "نسبة إشغال القاعات",
        "it_equipment_status_pct":     "% تجهيزات الإعلامية",
        "equipment_availability_rate": "توفر التجهيزات",
        "lab_availability_rate":       "توفر المخابر",
        "library_capacity_used_pct":   "إشغال المكتبة",
        "ongoing_works":               "الأشغال الجارية",
        "maintenance_requests":        "طلبات الصيانة",
        "resolved_requests":           "الطلبات المحلولة",
        "active_national_agreements":      "الاتفاقيات الوطنية النشطة",
        "active_international_agreements": "الاتفاقيات الدولية النشطة",
        "incoming_students":   "الطلاب الوافدون",
        "outgoing_students":   "الطلاب المتنقلون",
        "erasmus_partnerships": "شراكات إيراسموس+",
        "joint_programs":       "البرامج المشتركة",
        "industry_partnerships": "شراكات المؤسسات",
        "international_projects": "المشاريع الدولية",
        "graduates_total":              "إجمالي الخريجين",
        "employed_within_6months":      "موظفون في 6 أشهر",
        "employed_within_12months":     "موظفون في 12 شهرا",
        "employability_rate_6m":        "نسبة التشغيل 6أش",
        "employability_rate_12m":       "نسبة التشغيل 12أش",
        "avg_months_to_employment":     "متوسط التوظيف",
        "national_employment_pct":      "% التوظيف الوطني",
        "international_employment_pct": "% التوظيف الدولي",
        "self_employed_pct":            "% المقاولون الذاتيون",
        "energy_consumption_kwh":   "استهلاك الطاقة",
        "carbon_footprint_tons":    "البصمة الكربونية",
        "recycling_rate":           "نسبة إعادة التدوير",
        "green_spaces_sqm":         "المساحات الخضراء",
        "sustainable_mobility_pct": "التنقل المستدام",
        "accessibility_score":      "نقاط إمكانية الوصول",
        "waste_produced_tons":      "النفايات المنتجة",
        "water_consumption_m3":     "استهلاك المياه",
        "publications_count":            "المنشورات",
        "active_projects":               "المشاريع النشطة",
        "funding_secured_tnd":           "التمويلات المحصلة",
        "phd_students":                  "طلبة الدكتوراه",
        "patents_filed":                 "البراءات المودعة",
        "international_collaborations":  "التعاون الدولي",
        "national_collaborations":       "التعاون الوطني",
        "conferences_attended":          "المؤتمرات",
        "footer": "أُنشئ تلقائياً بواسطة منصة UCAR Intelligence · {n} مجالات مغطاة · سري · للاستخدام الداخلي فقط",
        "normal":    "✅ طبيعي",
        "attention": "⚠️ تنبيه",
        "critical":  "🔴 حرج",
        "unit_pct":    "%",
        "unit_tnd":    " د.ت",
        "unit_kwh":    " كيلوواط",
        "unit_tons":   " طن",
        "unit_sqm":    " م²",
        "unit_m3":     " م³",
        "unit_grade":  "/20",
        "unit_months": " أشهر",
        "unit_hrs":    "س/أسبوع",
    },
}


def _t(key: str, lang: str) -> str:
    labels = _L.get(lang, _L["fr"])
    text = labels.get(key, _L["fr"].get(key, key))
    return _ar(text) if lang == "ar" else text


def _status(value: float, warn: float, ok: float, lang: str) -> str:
    if value >= ok:
        return _t("normal", lang)
    elif value >= warn:
        return _t("attention", lang)
    return _t("critical", lang)


def _status_inv(value: float, warn: float, crit: float, lang: str) -> str:
    if value <= warn:
        return _t("normal", lang)
    elif value <= crit:
        return _t("attention", lang)
    return _t("critical", lang)


# ─── PDF Report ────────────────────────────────────────────────────────────────

def generate_pdf_report(
    institution: dict,
    period: str,
    academic: dict,
    finance: dict,
    hr: dict,
    alerts: list,
    ai_summary: str,
    infrastructure: dict = None,
    partnership: dict = None,
    employment: dict = None,
    esg: dict = None,
    research: dict = None,
    lang: str = "fr",
) -> bytes:
    is_ar = lang == "ar"
    font_name = _ARABIC_FONT if is_ar else "Helvetica"
    font_bold = (_ARABIC_FONT if is_ar else "Helvetica-Bold")
    align = TA_RIGHT if is_ar else TA_LEFT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    title_style    = ParagraphStyle('Title',    fontSize=20, textColor=UCAR_BLUE, spaceAfter=6,  fontName=font_bold,  alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', fontSize=11, textColor=HexColor('#555555'),       spaceAfter=20,       alignment=TA_CENTER, fontName=font_name)
    section_style  = ParagraphStyle('Section',  fontSize=13, textColor=UCAR_BLUE, spaceBefore=16, spaceAfter=8,        fontName=font_bold)
    section_green  = ParagraphStyle('SectionG', fontSize=13, textColor=ESG_GREEN,  spaceBefore=16, spaceAfter=8,        fontName=font_bold)
    section_purple = ParagraphStyle('SectionP', fontSize=13, textColor=RESEARCH_PURPLE, spaceBefore=16, spaceAfter=8,  fontName=font_bold)
    body_style     = ParagraphStyle('Body',     fontSize=10, spaceAfter=6, leading=14, fontName=font_name, alignment=align)
    footer_style   = ParagraphStyle('Footer',   fontSize=8,  textColor=HexColor('#999999'), alignment=TA_CENTER, fontName=font_name)

    def T(key): return _t(key, lang)

    elements = []

    # Header
    elements.append(Paragraph(T("university"), title_style))
    elements.append(Paragraph(
        f"{T('report_title')} — {institution.get('name_fr', '')}",
        subtitle_style
    ))
    elements.append(Paragraph(
        f"{T('period_label')}: {period} | {T('generated')}: {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        subtitle_style
    ))
    elements.append(HRFlowable(width="100%", thickness=2, color=UCAR_BLUE))
    elements.append(Spacer(1, 0.5*cm))

    # Institution info block
    info_data = [
        [_ar(T("code")) if is_ar else T("code"), institution.get('code', 'N/A'),
         _ar(T("governorate")) if is_ar else T("governorate"), institution.get('governorate', 'N/A')],
        [_ar(T("director")) if is_ar else T("director"), institution.get('director_name', 'N/A'),
         _ar(T("period_covered")) if is_ar else T("period_covered"), period],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), UCAR_LIGHT),
        ('BACKGROUND', (2, 0), (2, -1), UCAR_LIGHT),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTNAME', (0, 0), (0, -1), font_bold),
        ('FONTNAME', (2, 0), (2, -1), font_bold),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*cm))

    # AI Summary
    elements.append(Paragraph(T("exec_summary"), section_style))
    elements.append(Paragraph(ai_summary.replace('\n', '<br/>'), body_style))
    elements.append(Spacer(1, 0.4*cm))

    u = _L[lang]

    def pct(v):  return f"{v}{u['unit_pct']}"
    def tnd(v):  return f"{float(v or 0):,.0f}{u['unit_tnd']}"
    def grade(v): return f"{v}{u['unit_grade']}"

    # ── Academic ──
    elements.append(HRFlowable(width="100%", thickness=1, color=UCAR_LIGHT))
    elements.append(Paragraph(T("academic"), section_style))
    elements.append(_build_table([
        [T("indicator"), T("value"), T("status")],
        [T("total_enrolled"),   str(academic.get('total_enrolled', 'N/A')), ''],
        [T("total_passed"),     str(academic.get('total_passed', 'N/A')), ''],
        [T("success_rate"),     pct(academic.get('success_rate', 'N/A')),   _status(float(academic.get('success_rate') or 0), 70, 80, lang)],
        [T("dropout_rate"),     pct(academic.get('dropout_rate', 'N/A')),   _status_inv(float(academic.get('dropout_rate') or 0), 8, 15, lang)],
        [T("attendance_rate"),  pct(academic.get('attendance_rate', 'N/A')), _status(float(academic.get('attendance_rate') or 0), 75, 85, lang)],
        [T("repetition_rate"),  pct(academic.get('repetition_rate', 'N/A')), _status_inv(float(academic.get('repetition_rate') or 0), 10, 18, lang)],
        [T("avg_grade"),        grade(academic.get('avg_grade', 'N/A')), ''],
    ], font_name))
    elements.append(Spacer(1, 0.4*cm))

    # ── Finance ──
    elements.append(Paragraph(T("finance"), section_style))
    elements.append(_build_table([
        [T("indicator"), T("value"), T("status")],
        [T("allocated_budget"),      tnd(finance.get('allocated_budget')),      ''],
        [T("consumed_budget"),       tnd(finance.get('consumed_budget')),        ''],
        [T("budget_execution_rate"), pct(finance.get('budget_execution_rate', 'N/A')), _status(float(finance.get('budget_execution_rate') or 0), 60, 80, lang)],
        [T("cost_per_student"),      tnd(finance.get('cost_per_student')),       ''],
        [T("staff_budget_pct"),      pct(finance.get('staff_budget_pct', 'N/A')), ''],
        [T("infrastructure_budget_pct"), pct(finance.get('infrastructure_budget_pct', 'N/A')), ''],
        [T("research_budget_pct"),   pct(finance.get('research_budget_pct', 'N/A')), ''],
    ], font_name))
    elements.append(Spacer(1, 0.4*cm))

    # ── HR ──
    elements.append(Paragraph(T("hr"), section_style))
    elements.append(_build_table([
        [T("indicator"), T("value"), T("status")],
        [T("total_teaching_staff"),    str(hr.get('total_teaching_staff', 'N/A')), ''],
        [T("total_admin_staff"),       str(hr.get('total_admin_staff', 'N/A')),    ''],
        [T("absenteeism_rate"),        pct(hr.get('absenteeism_rate', 'N/A')),     _status_inv(float(hr.get('absenteeism_rate') or 0), 9, 15, lang)],
        [T("avg_teaching_load_hours"), f"{hr.get('avg_teaching_load_hours', 'N/A')}{u['unit_hrs']}", _status_inv(float(hr.get('avg_teaching_load_hours') or 0), 22, 28, lang)],
        [T("staff_turnover_rate"),     pct(hr.get('staff_turnover_rate', 'N/A')),  _status_inv(float(hr.get('staff_turnover_rate') or 0), 8, 15, lang)],
        [T("training_completion_rate"), pct(hr.get('training_completion_rate', 'N/A')), _status(float(hr.get('training_completion_rate') or 0), 60, 75, lang)],
        [T("permanent_staff_pct"),     pct(hr.get('permanent_staff_pct', 'N/A')), ''],
        [T("contract_staff_pct"),      pct(hr.get('contract_staff_pct', 'N/A')),  ''],
    ], font_name))
    elements.append(Spacer(1, 0.4*cm))

    # ── Infrastructure ──
    if infrastructure:
        elements.append(Paragraph(T("infrastructure"), section_style))
        elements.append(_build_table([
            [T("indicator"), T("value"), T("status")],
            [T("classroom_occupancy_rate"),    pct(infrastructure.get('classroom_occupancy_rate', 'N/A')),    _status(float(infrastructure.get('classroom_occupancy_rate') or 0), 60, 80, lang)],
            [T("it_equipment_status_pct"),     pct(infrastructure.get('it_equipment_status_pct', 'N/A')),     _status(float(infrastructure.get('it_equipment_status_pct') or 0), 70, 85, lang)],
            [T("equipment_availability_rate"), pct(infrastructure.get('equipment_availability_rate', 'N/A')), _status(float(infrastructure.get('equipment_availability_rate') or 0), 75, 88, lang)],
            [T("lab_availability_rate"),       pct(infrastructure.get('lab_availability_rate', 'N/A')),       _status(float(infrastructure.get('lab_availability_rate') or 0), 70, 85, lang)],
            [T("library_capacity_used_pct"),   pct(infrastructure.get('library_capacity_used_pct', 'N/A')),  ''],
            [T("ongoing_works"),    str(infrastructure.get('ongoing_works', 'N/A')),    ''],
            [T("maintenance_requests"), str(infrastructure.get('maintenance_requests', 'N/A')), ''],
            [T("resolved_requests"),    str(infrastructure.get('resolved_requests', 'N/A')),    ''],
        ], font_name))
        elements.append(Spacer(1, 0.4*cm))

    # ── Partnership ──
    if partnership:
        elements.append(Paragraph(T("partnership"), section_style))
        elements.append(_build_table([
            [T("indicator"), T("value"), T("status")],
            [T("active_national_agreements"),      str(partnership.get('active_national_agreements', 'N/A')),      ''],
            [T("active_international_agreements"), str(partnership.get('active_international_agreements', 'N/A')), ''],
            [T("incoming_students"),    str(partnership.get('incoming_students', 'N/A')),   ''],
            [T("outgoing_students"),    str(partnership.get('outgoing_students', 'N/A')),   ''],
            [T("erasmus_partnerships"), str(partnership.get('erasmus_partnerships', 'N/A')), ''],
            [T("joint_programs"),       str(partnership.get('joint_programs', 'N/A')),       ''],
            [T("industry_partnerships"), str(partnership.get('industry_partnerships', 'N/A')), ''],
            [T("international_projects"), str(partnership.get('international_projects', 'N/A')), ''],
        ], font_name))
        elements.append(Spacer(1, 0.4*cm))

    # ── Employment ──
    if employment:
        elements.append(Paragraph(T("employment"), section_style))
        elements.append(_build_table([
            [T("indicator"), T("value"), T("status")],
            [T("graduates_total"),              str(employment.get('graduates_total', 'N/A')),          ''],
            [T("employed_within_6months"),      str(employment.get('employed_within_6months', 'N/A')),  ''],
            [T("employed_within_12months"),     str(employment.get('employed_within_12months', 'N/A')), ''],
            [T("employability_rate_6m"),        pct(employment.get('employability_rate_6m', 'N/A')),    _status(float(employment.get('employability_rate_6m') or 0), 35, 55, lang)],
            [T("employability_rate_12m"),       pct(employment.get('employability_rate_12m', 'N/A')),   _status(float(employment.get('employability_rate_12m') or 0), 50, 70, lang)],
            [T("avg_months_to_employment"),     f"{employment.get('avg_months_to_employment', 'N/A')}{u['unit_months']}", _status_inv(float(employment.get('avg_months_to_employment') or 0), 8, 14, lang)],
            [T("national_employment_pct"),      pct(employment.get('national_employment_pct', 'N/A')),  ''],
            [T("international_employment_pct"), pct(employment.get('international_employment_pct', 'N/A')), ''],
            [T("self_employed_pct"),            pct(employment.get('self_employed_pct', 'N/A')),         ''],
        ], font_name))
        elements.append(Spacer(1, 0.4*cm))

    # ── ESG ──
    if esg:
        elements.append(Paragraph(T("esg"), section_green))
        elements.append(_build_table_colored([
            [T("indicator"), T("value"), T("status")],
            [T("energy_consumption_kwh"),   f"{float(esg.get('energy_consumption_kwh') or 0):,.0f}{u['unit_kwh']}", ''],
            [T("carbon_footprint_tons"),    f"{esg.get('carbon_footprint_tons', 'N/A')}{u['unit_tons']}",           _status_inv(float(esg.get('carbon_footprint_tons') or 0), 80, 150, lang)],
            [T("recycling_rate"),           pct(esg.get('recycling_rate', 'N/A')),                                  _status(float(esg.get('recycling_rate') or 0), 25, 35, lang)],
            [T("green_spaces_sqm"),         f"{esg.get('green_spaces_sqm', 'N/A')}{u['unit_sqm']}",                ''],
            [T("sustainable_mobility_pct"), pct(esg.get('sustainable_mobility_pct', 'N/A')),                        _status(float(esg.get('sustainable_mobility_pct') or 0), 30, 45, lang)],
            [T("accessibility_score"),      f"{esg.get('accessibility_score', 'N/A')}/100",                         _status(float(esg.get('accessibility_score') or 0), 55, 70, lang)],
            [T("waste_produced_tons"),      f"{esg.get('waste_produced_tons', 'N/A')}{u['unit_tons']}",             ''],
            [T("water_consumption_m3"),     f"{float(esg.get('water_consumption_m3') or 0):,.0f}{u['unit_m3']}",   ''],
        ], ESG_GREEN, ESG_LIGHT, font_name))
        elements.append(Spacer(1, 0.4*cm))

    # ── Research ──
    if research:
        elements.append(Paragraph(T("research"), section_purple))
        elements.append(_build_table_colored([
            [T("indicator"), T("value"), T("status")],
            [T("publications_count"),           str(research.get('publications_count', 'N/A')),          ''],
            [T("active_projects"),              str(research.get('active_projects', 'N/A')),             ''],
            [T("funding_secured_tnd"),          f"{float(research.get('funding_secured_tnd') or 0):,.0f}{u['unit_tnd']}", ''],
            [T("phd_students"),                 str(research.get('phd_students', 'N/A')),                ''],
            [T("patents_filed"),                str(research.get('patents_filed', 'N/A')),               ''],
            [T("international_collaborations"), str(research.get('international_collaborations', 'N/A')), ''],
            [T("national_collaborations"),      str(research.get('national_collaborations', 'N/A')),      ''],
            [T("conferences_attended"),         str(research.get('conferences_attended', 'N/A')),         ''],
        ], RESEARCH_PURPLE, RESEARCH_LIGHT, font_name))
        elements.append(Spacer(1, 0.4*cm))

    # ── Alerts ──
    if alerts:
        elements.append(Paragraph(T("alerts"), section_style))
        alert_data = [[T("severity"), T("domain"), T("title")]]
        for alert in alerts[:5]:
            alert_data.append([
                alert.get('severity', '').upper(),
                alert.get('domain', ''),
                alert.get('title', '')[:60],
            ])
        elements.append(_build_alert_table(alert_data, font_name))
        elements.append(Spacer(1, 0.4*cm))

    # Footer
    elements.append(HRFlowable(width="100%", thickness=1, color=UCAR_BLUE))
    elements.append(Spacer(1, 0.2*cm))
    domains_covered = sum([bool(academic), bool(finance), bool(hr),
                           bool(infrastructure), bool(partnership), bool(employment),
                           bool(esg), bool(research)])
    footer_text = _L[lang]["footer"].format(n=domains_covered)
    if is_ar:
        footer_text = _ar(footer_text)
    elements.append(Paragraph(footer_text, footer_style))

    doc.build(elements)
    return buffer.getvalue()


# ─── Excel Report ──────────────────────────────────────────────────────────────

def generate_excel_report(
    institution: dict,
    period: str,
    academic: dict,
    finance: dict,
    hr: dict,
    infrastructure: dict = None,
    partnership: dict = None,
    employment: dict = None,
    esg: dict = None,
    research: dict = None,
    lang: str = "fr",
) -> bytes:
    def T(key): return _L.get(lang, _L["fr"]).get(key, _L["fr"].get(key, key))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = T("report_title")[:31]

    header_font    = Font(bold=True, color="FFFFFF", size=12)
    header_fill    = PatternFill(start_color="1a3c6e", end_color="1a3c6e", fill_type="solid")
    section_font   = Font(bold=True, color="1a3c6e", size=11)
    section_fill   = PatternFill(start_color="e8f0fa", end_color="e8f0fa", fill_type="solid")
    esg_font       = Font(bold=True, color="1e8449", size=11)
    esg_fill       = PatternFill(start_color="d5f5e3", end_color="d5f5e3", fill_type="solid")
    research_font  = Font(bold=True, color="6c3483", size=11)
    research_fill  = PatternFill(start_color="f0e6f6", end_color="f0e6f6", fill_type="solid")
    is_ar = lang == "ar"
    h_align = 'right' if is_ar else 'center'
    center  = Alignment(horizontal=h_align, vertical='center')
    thin    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"{T('university')} — {institution.get('name_fr', '')}"
    ws['A1'].font = Font(bold=True, size=14, color="1a3c6e")
    ws['A1'].alignment = center

    ws.merge_cells('A2:C2')
    ws['A2'] = f"{T('report_title')} · {T('period_label')}: {period} · {T('generated')} {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = center
    ws['A2'].font = Font(italic=True, size=10, color="888888")

    row = 4

    def write_section(title, data_dict, sf=section_font, sfill=section_fill):
        nonlocal row
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = sf
        ws[f'A{row}'].fill = sfill
        ws[f'A{row}'].alignment = center
        row += 1
        for label, value in data_dict.items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = thin
            ws[f'B{row}'].border = thin
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        row += 1

    u = _L[lang]

    write_section(T("academic"), {
        T("total_enrolled"):   academic.get('total_enrolled', 'N/A'),
        T("total_passed"):     academic.get('total_passed', 'N/A'),
        T("success_rate"):     float(academic.get('success_rate') or 0),
        T("dropout_rate"):     float(academic.get('dropout_rate') or 0),
        T("attendance_rate"):  float(academic.get('attendance_rate') or 0),
        T("repetition_rate"):  float(academic.get('repetition_rate') or 0),
        T("avg_grade"):        float(academic.get('avg_grade') or 0),
    })

    write_section(T("finance"), {
        T("allocated_budget"):         float(finance.get('allocated_budget') or 0),
        T("consumed_budget"):          float(finance.get('consumed_budget') or 0),
        T("budget_execution_rate"):    float(finance.get('budget_execution_rate') or 0),
        T("cost_per_student"):         float(finance.get('cost_per_student') or 0),
        T("staff_budget_pct"):         float(finance.get('staff_budget_pct') or 0),
        T("infrastructure_budget_pct"): float(finance.get('infrastructure_budget_pct') or 0),
        T("research_budget_pct"):      float(finance.get('research_budget_pct') or 0),
    })

    write_section(T("hr"), {
        T("total_teaching_staff"):    hr.get('total_teaching_staff', 'N/A'),
        T("total_admin_staff"):       hr.get('total_admin_staff', 'N/A'),
        T("absenteeism_rate"):        float(hr.get('absenteeism_rate') or 0),
        T("avg_teaching_load_hours"): float(hr.get('avg_teaching_load_hours') or 0),
        T("staff_turnover_rate"):     float(hr.get('staff_turnover_rate') or 0),
        T("training_completion_rate"): float(hr.get('training_completion_rate') or 0),
        T("permanent_staff_pct"):     float(hr.get('permanent_staff_pct') or 0),
        T("contract_staff_pct"):      float(hr.get('contract_staff_pct') or 0),
    })

    if infrastructure:
        write_section(T("infrastructure"), {
            T("classroom_occupancy_rate"):    float(infrastructure.get('classroom_occupancy_rate') or 0),
            T("it_equipment_status_pct"):     float(infrastructure.get('it_equipment_status_pct') or 0),
            T("equipment_availability_rate"): float(infrastructure.get('equipment_availability_rate') or 0),
            T("lab_availability_rate"):       float(infrastructure.get('lab_availability_rate') or 0),
            T("library_capacity_used_pct"):   float(infrastructure.get('library_capacity_used_pct') or 0),
            T("ongoing_works"):        infrastructure.get('ongoing_works', 0),
            T("maintenance_requests"): infrastructure.get('maintenance_requests', 0),
            T("resolved_requests"):    infrastructure.get('resolved_requests', 0),
        })

    if partnership:
        write_section(T("partnership"), {
            T("active_national_agreements"):      partnership.get('active_national_agreements', 0),
            T("active_international_agreements"): partnership.get('active_international_agreements', 0),
            T("incoming_students"):    partnership.get('incoming_students', 0),
            T("outgoing_students"):    partnership.get('outgoing_students', 0),
            T("erasmus_partnerships"): partnership.get('erasmus_partnerships', 0),
            T("joint_programs"):       partnership.get('joint_programs', 0),
            T("industry_partnerships"): partnership.get('industry_partnerships', 0),
            T("international_projects"): partnership.get('international_projects', 0),
        })

    if employment:
        write_section(T("employment"), {
            T("graduates_total"):              employment.get('graduates_total', 0),
            T("employed_within_6months"):      employment.get('employed_within_6months', 0),
            T("employed_within_12months"):     employment.get('employed_within_12months', 0),
            T("employability_rate_6m"):        float(employment.get('employability_rate_6m') or 0),
            T("employability_rate_12m"):       float(employment.get('employability_rate_12m') or 0),
            T("avg_months_to_employment"):     float(employment.get('avg_months_to_employment') or 0),
            T("national_employment_pct"):      float(employment.get('national_employment_pct') or 0),
            T("international_employment_pct"): float(employment.get('international_employment_pct') or 0),
            T("self_employed_pct"):            float(employment.get('self_employed_pct') or 0),
        })

    if esg:
        write_section(T("esg"), {
            T("energy_consumption_kwh"):   float(esg.get('energy_consumption_kwh') or 0),
            T("carbon_footprint_tons"):    float(esg.get('carbon_footprint_tons') or 0),
            T("recycling_rate"):           float(esg.get('recycling_rate') or 0),
            T("green_spaces_sqm"):         esg.get('green_spaces_sqm', 0),
            T("sustainable_mobility_pct"): float(esg.get('sustainable_mobility_pct') or 0),
            T("accessibility_score"):      float(esg.get('accessibility_score') or 0),
            T("waste_produced_tons"):      float(esg.get('waste_produced_tons') or 0),
            T("water_consumption_m3"):     float(esg.get('water_consumption_m3') or 0),
        }, sf=esg_font, sfill=esg_fill)

    if research:
        write_section(T("research"), {
            T("publications_count"):            research.get('publications_count', 0),
            T("active_projects"):               research.get('active_projects', 0),
            T("funding_secured_tnd"):           float(research.get('funding_secured_tnd') or 0),
            T("phd_students"):                  research.get('phd_students', 0),
            T("patents_filed"):                 research.get('patents_filed', 0),
            T("international_collaborations"):  research.get('international_collaborations', 0),
            T("national_collaborations"):       research.get('national_collaborations', 0),
            T("conferences_attended"):          research.get('conferences_attended', 0),
        }, sf=research_font, sfill=research_fill)

    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].width = 38

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─── Table builders ────────────────────────────────────────────────────────────

def _build_table(data: list, font_name: str = "Helvetica"):
    t = Table(data, colWidths=[8*cm, 4*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), UCAR_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (2, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, UCAR_LIGHT]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    return t


def _build_table_colored(data: list, header_color, row_color, font_name: str = "Helvetica"):
    t = Table(data, colWidths=[8*cm, 4*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (2, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, row_color]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    return t


def _build_alert_table(data: list, font_name: str = "Helvetica"):
    t = Table(data, colWidths=[3*cm, 3*cm, 10*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), UCAR_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#fff5f5'), white]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    return t
